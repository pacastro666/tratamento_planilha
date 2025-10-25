import io
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from zipfile import BadZipFile

# ----------------------------
# Utilidades
# ----------------------------
def extract_numbers_from_text(text: str):
    if pd.isna(text):
        return []
    return [int(n) for n in re.findall(r"\d+", str(text))]

def build_number_to_consultant_map(df_reat10_raw: pd.DataFrame):
    # Espera-se:
    # Coluna A (index 0): CONSULTOR
    # Coluna B (index 1): PEDIDOS (sequências com números)
    consultor_col = df_reat10_raw.iloc[:, 0].astype(str)
    pedidos_col = df_reat10_raw.iloc[:, 1]

    num_to_consultants = {}
    for cons, pedidos in zip(consultor_col, pedidos_col):
        if pd.isna(pedidos):
            continue
        for n in extract_numbers_from_text(pedidos):
            s = num_to_consultants.setdefault(n, set())
            s.add(cons)
    return num_to_consultants

def create_or_update_extracted_sheet(wb: Workbook, sheet_source: str = "REAT-10", sheet_out: str = "EXTRAIDOS_REAT10"):
    # Ler REAT-10 para mapear números -> consultores
    ws_src = wb[sheet_source]
    last_row_src = ws_src.max_row

    # Carregar REAT-10 em DataFrame bruto (sem header)
    # Como estamos em memória com openpyxl, vamos montar um DF manualmente
    data = []
    for r in ws_src.iter_rows(values_only=True):
        data.append(list(r))
    df_raw = pd.DataFrame(data)

    # Mapear número -> consultores
    num_to_cons = build_number_to_consultant_map(df_raw)
    # Números únicos em ordem de primeira aparição
    unique_numbers = list(num_to_cons.keys())

    # (Re)criar aba de saída
    if sheet_out in wb.sheetnames:
        ws_out = wb[sheet_out]
        wb.remove(ws_out)
    ws_out = wb.create_sheet(title=sheet_out)

    # Cabeçalhos
    ws_out.cell(row=1, column=1, value="NUMERO")
    ws_out.cell(row=1, column=2, value="ENCONTRADO_REAT10")
    ws_out.cell(row=1, column=3, value="OCORRENCIAS_REAT10")
    ws_out.cell(row=1, column=4, value="CONSULTORES")

    # Intervalo de busca para a fórmula
    lookup_range = f"'{sheet_source}'!$B$1:$B${last_row_src}"

    # Preencher linhas
    for i, num in enumerate(unique_numbers, start=2):
        ws_out.cell(row=i, column=1, value=num)

        # Fórmula que considera números dentro de sequências com '-'
        # =SUMPRODUCT(--ISNUMBER(SEARCH("-"&A2&"-","-"&SUBSTITUTE('REAT-10'!$B$1:$B$28," ","")&"-")))>0
        formula = f'=SUMPRODUCT(--ISNUMBER(SEARCH("-"&A{i}&"-","-"&SUBSTITUTE({lookup_range}," ","")&"-")))>0'
        ws_out.cell(row=i, column=2, value=formula)

        consultants = sorted(
            x for x in num_to_cons[num]
            if x and x.lower() not in ("nan", "consultor")
        )
        ws_out.cell(row=i, column=3, value=len(consultants))
        ws_out.cell(row=i, column=4, value="; ".join(consultants))

    # Estética básica
    ws_out.auto_filter.ref = f"A1:D{len(unique_numbers)+1}"
    ws_out.column_dimensions['A'].width = 16
    ws_out.column_dimensions['B'].width = 22
    ws_out.column_dimensions['C'].width = 22
    ws_out.column_dimensions['D'].width = 50

    return unique_numbers, last_row_src  # para exibição no app

def get_dataframe_from_sheet(wb: Workbook, sheet_name: str):
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    # Detectar header da primeira linha
    if len(data) == 0:
        return pd.DataFrame()
    
    # Se primeira linha parece cabeçalho (strings), usar como header
    header = data[0]
    df = pd.DataFrame(data[1:], columns=header)
    
    # Limpar tipos de dados para evitar problemas com PyArrow
    for col in df.columns:
        # Converter colunas com tipos mistos para string
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str)
            # Substituir 'None' e 'nan' por string vazia
            df[col] = df[col].replace(['None', 'nan', 'NaN'], '')
    
    return df

def load_excel_file(file_content):
    """
    Carrega um arquivo Excel com tratamento de erro
    """
    try:
        bio = io.BytesIO(file_content)
        wb = load_workbook(bio)
        return wb, None
    except BadZipFile:
        return None, "❌ Erro: O arquivo não é um Excel válido (.xlsx). Verifique se o arquivo não está corrompido."
    except Exception as e:
        return None, f"❌ Erro ao carregar arquivo: {str(e)}"

# ----------------------------
# Streamlit App
# ----------------------------
st.set_page_config(page_title="Extrator de Números Excel", layout="wide")
st.title("Extrator de Números de Planilhas Excel")

# Informações sobre tipos de arquivo
with st.expander("ℹ️ Informações sobre tipos de arquivo"):
    st.markdown("""
    **Tipos de arquivo suportados:**
    - ✅ **.xlsx** - Excel moderno (recomendado)
    
    **Tipos NÃO suportados:**
    - ❌ **.xls** - Excel antigo (versão 97-2003)
    - ❌ **.csv** - Arquivos de texto
    - ❌ **.ods** - OpenDocument
    
    **Se você tem um arquivo .xls:**
    1. Abra no Excel
    2. Salve como .xlsx (Arquivo → Salvar Como → Escolher formato .xlsx)
    3. Faça upload do novo arquivo
    """)

with st.sidebar:
    st.header("Importar Planilha")
    uploaded = st.file_uploader("Selecione o arquivo Excel", type=["xlsx"], help="Apenas arquivos .xlsx são suportados")
    
    process_btn = st.button("Processar e atualizar aba")

# Estado
if "workbook_bytes" not in st.session_state:
    st.session_state.workbook_bytes = None
if "report" not in st.session_state:
    st.session_state.report = None
if "selected_sheet" not in st.session_state:
    st.session_state.selected_sheet = None
if "output_sheet_name" not in st.session_state:
    st.session_state.output_sheet_name = "EXTRAIDOS_REAT10"

tab1, tab2, tab3 = st.tabs(["📊 Dados Originais", "📈 Dados Processados", "💾 Download"])

wb = None

if uploaded is not None:
    # Carregar workbook na memória
    content = uploaded.read()
    st.session_state.workbook_bytes = content
    wb, error_msg = load_excel_file(content)
    
    if wb is None:
        st.error(error_msg)
        st.info("💡 **Dicas para resolver:**")
        st.info("• Verifique se o arquivo é realmente um Excel (.xlsx)")
        st.info("• Tente salvar o arquivo novamente no Excel")
        st.info("• Verifique se o arquivo não está corrompido")
        st.info("• Certifique-se de que não é um arquivo .xls (versão antiga)")
        st.stop()  # Para a execução se houver erro
    
    # Configurações após carregamento bem-sucedido
    with st.sidebar:
        st.subheader("Configurações")
        
        # Selecionar aba
        available_sheets = wb.sheetnames
        st.session_state.selected_sheet = st.selectbox(
            "Selecione a aba para processar:",
            available_sheets,
            help="Escolha qual aba contém os dados que você quer processar"
        )
        
        # Nome da aba de saída
        st.session_state.output_sheet_name = st.text_input(
            "Nome da aba de saída:",
            value=st.session_state.output_sheet_name,
            help="Nome da aba que será criada com os resultados"
        )

    with tab1:
        if st.session_state.selected_sheet:
            st.subheader(f"Dados originais da aba '{st.session_state.selected_sheet}'")
            if st.session_state.selected_sheet in wb.sheetnames:
                df_sheet = get_dataframe_from_sheet(wb, st.session_state.selected_sheet)
                st.dataframe(df_sheet, width='stretch', height=500)
            else:
                st.warning(f"A aba '{st.session_state.selected_sheet}' não foi encontrada no arquivo.")
        else:
            st.info("Selecione uma aba na barra lateral para visualizar os dados.")

    if process_btn:
        if wb is None:
            st.error("Erro ao carregar workbook.")
        elif not st.session_state.selected_sheet:
            st.error("Selecione uma aba para processar.")
        elif st.session_state.selected_sheet not in wb.sheetnames:
            st.error(f"A aba '{st.session_state.selected_sheet}' não foi encontrada. Verifique o arquivo.")
        else:
            # Processar e atualizar aba
            unique_numbers, last_row_src = create_or_update_extracted_sheet(wb, st.session_state.selected_sheet, st.session_state.output_sheet_name)

            # Salvar workbook modificado em memória
            out_buf = io.BytesIO()
            wb.save(out_buf)
            st.session_state.workbook_bytes = out_buf.getvalue()

            st.session_state.report = {
                "unique_count": len(unique_numbers),
                "last_row_src": last_row_src,
                "source_sheet": st.session_state.selected_sheet,
                "output_sheet": st.session_state.output_sheet_name
            }
            st.success(f"Aba '{st.session_state.output_sheet_name}' atualizada. Números únicos: {len(unique_numbers)} | Última linha da '{st.session_state.selected_sheet}': {last_row_src}")

    # Visualização da aba gerada (se já foi processada nesta sessão)
    with tab2:
        output_sheet = st.session_state.output_sheet_name
        source_sheet = st.session_state.selected_sheet if st.session_state.selected_sheet else "aba selecionada"
        st.subheader(f"Dados processados da aba '{source_sheet}' → '{output_sheet}'")
        if wb is not None and output_sheet in wb.sheetnames:
            df_out = get_dataframe_from_sheet(wb, output_sheet)
            st.dataframe(df_out, width='stretch', height=500)
        else:
            st.info("Ainda não há aba processada para visualizar. Clique em 'Processar e atualizar aba'.")

    # Download do arquivo atualizado
    with tab3:
        st.subheader("Baixar arquivo Excel modificado")
        if st.session_state.workbook_bytes:
            st.download_button(
                label="Baixar .xlsx atualizado",
                data=st.session_state.workbook_bytes,
                file_name="Resultados2025_atualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if st.session_state.report:
                report = st.session_state.report
                st.caption(f"Números únicos: {report['unique_count']} | Última linha {report['source_sheet']}: {report['last_row_src']}")
                st.caption(f"Aba processada: {report['source_sheet']} → {report['output_sheet']}")
        else:
            st.info("Após processar, o arquivo atualizado ficará disponível para download aqui.")
else:
    st.info("Faça o upload de uma planilha .xlsx na barra lateral para começar.")
